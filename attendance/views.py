from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from .models import Student, Attendance, Department, Section
import qrcode
from io import BytesIO
import base64
import openpyxl
from datetime import datetime, date
from openpyxl.styles import Font, Alignment
import json

def index(request):
    return render(request, 'index.html')

def qrscanner(request):
    """View function to render the QR scanner page"""
    return render(request, 'qrscanner.html')

def generate_qr(request):
    """View function to render the QR generator page"""
    return render(request, 'qrgenerator.html')

def registration_success(request):
    return render(request, 'registration_success.html')

def scan_qr(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            lrn = data.get('lrn')
            name_parts = data.get('name', '').split()  # Split the full name
            class_info = data.get('class', '').split('&')  # Split grade and section
            action = data.get('action', 'time_in')
            parent_contact = data.get('parentContact')

            if not lrn:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Student LRN is required'
                })

            try:
                student = Student.objects.get(lrn=lrn)
                
                # Update student information if it has changed
                updated = False
                
                if len(name_parts) >= 2:
                    new_first_name = ' '.join(name_parts[:-1])  # All except last name
                    new_last_name = name_parts[-1]  # Last name
                    if student.first_name != new_first_name or student.last_name != new_last_name:
                        student.first_name = new_first_name
                        student.last_name = new_last_name
                        updated = True
                
                if len(class_info) >= 2:
                    section_name = class_info[1].strip()
                    grade_level = class_info[0].strip()
                    # Get or create department (assume JHS for Grade 7-10, SHS for Grade 11-12)
                    dept_name = 'SHS' if any(str(i) in grade_level for i in [11, 12]) else 'JHS'
                    department = Department.objects.get_or_create(name=dept_name)[0]
                    section = Section.objects.get_or_create(
                        name=section_name,
                        department=department
                    )[0]
                    if student.section != section:
                        student.section = section
                        updated = True
                
                if parent_contact and student.parent_mobile != parent_contact:
                    student.parent_mobile = parent_contact
                    updated = True
                
                if updated:
                    student.save()
                
                # Get current time in Manila timezone
                manila_tz = timezone.pytz.timezone('Asia/Manila')
                current_datetime = timezone.now().astimezone(manila_tz)
                current_date = current_datetime.date()
                current_time = current_datetime.time()
                
                # Get subject from the student's section (for now, just get any subject)
                current_subject = student.section.subjects.first()
                
                if not current_subject:
                    return JsonResponse({
                        'success': False,
                        'message': 'No subject assigned to this section'
                    })
                
                if action == 'time_in':
                    # Check if anyone has already attended this subject today (excluding this student)
                    first_attendance = Attendance.objects.filter(
                        subject=current_subject,
                        date=current_date
                    ).exclude(
                        student=student  # Exclude current student's attendance
                    ).order_by('time_in').first()

                    if first_attendance and first_attendance.time_in:
                        # Calculate minutes from first attendance
                        # Get time component from first_attendance
                        first_time_in = first_attendance.time_in.time() if isinstance(first_attendance.time_in, datetime) else first_attendance.time_in
                        
                        time_diff = datetime.combine(date.today(), current_time) - datetime.combine(date.today(), first_time_in)
                        minutes_late = int(time_diff.total_seconds() / 60)
                        
                        if minutes_late > 30:
                            return JsonResponse({
                                'success': False,
                                'message': (
                                    f'Sorry, you are {minutes_late} minutes late.\n'
                                    f'Time-in is not allowed after 30 minutes from the first student\'s arrival.\n\n'
                                    f'First student arrived at: {first_attendance.time_in.strftime("%I:%M %p")}\n'
                                    f'Current time: {current_time.strftime("%I:%M %p")}'
                                )
                            })

                    # For time_in, always create new or update existing
                    attendance, _ = Attendance.objects.update_or_create(
                        student=student,
                        subject=current_subject,
                        date=current_date,
                        defaults={'time_in': current_time}
                    )
                
                elif action == 'time_out':
                    # For time_out, get existing record and update time_out
                    attendance = Attendance.objects.filter(
                        student=student,
                        subject=current_subject,
                        date=current_date
                    ).first()
                    
                    if attendance:
                        attendance.time_out = current_time
                        attendance.save()
                    else:
                        # Create new attendance with time_out if none exists
                        attendance = Attendance.objects.create(
                            student=student,
                            subject=current_subject,
                            date=current_date,
                            time_out=current_time
                        )
                
                # Status will be automatically set by the model's save method
                status = attendance.status

                # Send email notification
                try:
                    subject = f'Student {"Time In" if action == "time_in" else "Time Out"} Notification'
                    current_time_str = current_time.strftime('%I:%M %p')
                    current_date_str = current_date.strftime('%B %d, %Y')
                    
                    context = {
                        'student_name': student.get_full_name(),
                        'action': 'arrived at school' if action == 'time_in' else 'left school',
                        'time': current_time_str,
                        'date': current_date_str,
                        'status': status,
                        'subject': current_subject.name if current_subject else 'N/A'
                    }
                    
                    message = render_to_string('email/attendance_notification.html', context)
                    
                    # Send to student's email if available
                    recipients = []
                    if student.email:
                        recipients.append(student.email)
                    
                    # Send to parent's email if available
                    if student.parent_email:
                        recipients.append(student.parent_email)
                        
                    if recipients:
                        send_mail(
                            subject,
                            message,
                            settings.EMAIL_HOST_USER,
                            recipients,
                            html_message=message,
                            fail_silently=True
                        )
                except Exception as e:
                    print(f"Error sending email: {str(e)}")
                
                return JsonResponse({
                    'success': True,
                    'type': action,
                    'message': f'{"Time-in" if action == "time_in" else "Time-out"} recorded for {student.get_full_name()}',
                    'time': current_time.strftime('%I:%M %p'),
                    'status': status,
                    'report_date': current_date.strftime('%Y-%m-%d') if action == 'time_in' else None
                })
                
            except Student.DoesNotExist:
                # Create new student record
                try:
                    # Get current time
                    current_date = timezone.now().date()
                    current_time = timezone.now().time()
                    
                    # Parse grade level from class info
                    if len(class_info) >= 2:
                        section_name = class_info[1].strip()
                        grade_level = class_info[0].strip()
                        dept_name = 'SHS' if any(str(i) in grade_level for i in [11, 12]) else 'JHS'
                        
                        # Create department and section
                        department = Department.objects.get_or_create(name=dept_name)[0]
                        section = Section.objects.get_or_create(
                            name=section_name,
                            department=department
                        )[0]
                        
                        # Create student with auto-generated student ID
                        if len(name_parts) >= 2:
                            # Generate unique student ID based on department and count
                            dept_prefix = 'SHS' if dept_name == 'SHS' else 'JHS'
                            year = timezone.now().year % 100  # Get last 2 digits of year
                            
                            # Count existing students in the department for this year
                            existing_count = Student.objects.filter(
                                section__department=department,
                                student_id__startswith=f'{dept_prefix}{year}'
                            ).count()
                            
                            # Format: JHS23001 or SHS23001 (dept + year + 3-digit sequence)
                            student_id = f'{dept_prefix}{year}{(existing_count + 1):03d}'
                            
                            # Create student first
                            student = Student.objects.create(
                                lrn=lrn,
                                student_id=student_id,
                                first_name=' '.join(name_parts[:-1]),
                                last_name=name_parts[-1],
                                section=section,
                                parent_mobile=parent_contact
                            )
                            
                            # Generate QR code for student
                            qr_data = {
                                'lrn': student.lrn,
                                'name': student.get_full_name(),
                                'class': str(student.section),
                                'parentContact': student.parent_mobile
                            }
                            
                            # Create QR code
                            qr = qrcode.QRCode(
                                version=1,
                                error_correction=qrcode.constants.ERROR_CORRECT_L,
                                box_size=10,
                                border=4,
                            )
                            qr.add_data(json.dumps(qr_data))
                            qr.make(fit=True)
                            qr_image = qr.make_image(fill_color="black", back_color="white")
                            
                            # Save QR code to BytesIO
                            qr_io = BytesIO()
                            qr_image.save(qr_io, format='PNG')
                            student.qr_code.save(
                                f'qr_{student.student_id}.png',
                                BytesIO(qr_io.getvalue()),
                                save=True
                            )
                            
                            # Get the first subject in the section
                            current_subject = section.subjects.first()
                            
                            if not current_subject:
                                return JsonResponse({
                                    'success': False,
                                    'message': 'No subjects assigned to this section yet.'
                                })
                            
                            # Record initial attendance
                            attendance = Attendance.objects.create(
                                student=student,
                                subject=current_subject,
                                date=current_date,
                                time_in=current_time if action == 'time_in' else None,
                                time_out=current_time if action == 'time_out' else None,
                                status='PRESENT'  # Set default status
                            )
                            
                            registration_message = (
                                f"Student successfully registered!\n"
                                f"Student ID: {student.student_id}\n"
                                f"Name: {student.get_full_name()}\n"
                                f"Section: {student.section.name}\n"
                                f"Department: {student.section.department}"
                            )

                            # Get QR code URL for response
                            qr_url = student.qr_code.url if student.qr_code else None

                            return JsonResponse({
                                'success': True,
                                'type': action,
                                'message': registration_message,
                                'student_id': student.student_id,
                                'time': current_time.strftime('%I:%M %p'),
                                'status': None,
                                'report_date': current_date.strftime('%Y-%m-%d') if action == 'time_in' else None,
                                'is_new_registration': True,
                                'qr_code_url': qr_url  # Include QR code URL in response
                            })
                    
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'message': f'Error creating new student record: {str(e)}'
                    })
                    
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid student data format'
                })
                
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid QR code data format'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            })
    
    # Add CSRF token to the template context
    return render(request, 'qrscanner.html')

def export_excel(request):
    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Write headers
    headers = ['Student ID', 'Name', 'Section', 'Department', 'Date', 'Time In', 'Time Out', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Get all attendance records
    records = Attendance.objects.all().order_by('date', 'student__section', 'student__last_name')
    
    # Write data
    for row, record in enumerate(records, 2):
        ws.cell(row=row, column=1, value=record.student.student_id)
        ws.cell(row=row, column=2, value=record.student.get_full_name())
        ws.cell(row=row, column=3, value=record.student.section.name)
        ws.cell(row=row, column=4, value=record.student.section.department.get_name_display())
        ws.cell(row=row, column=5, value=record.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=6, value=record.time_in.strftime('%I:%M %p') if record.time_in else '')
        ws.cell(row=row, column=7, value=record.time_out.strftime('%I:%M %p') if record.time_out else '')
        ws.cell(row=row, column=8, value=record.status)
    
    # Set column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=attendance_report_{date.today().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response