from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ..models import Student, Attendance  # You'll need to create Attendance model if not exists

def export_excel(request):
    """Export attendance data to Excel"""
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Define styles
    header_style = {
        'font': Font(bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid'),
        'alignment': Alignment(horizontal='center')
    }

    # Write headers
    headers = ['LRN', 'Student Name', 'Grade & Section', 'Date', 'Time In', 'Time Out', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_style['font']
        cell.fill = header_style['fill']
        cell.alignment = header_style['alignment']

    try:
        # Get attendance records (modify this according to your model structure)
        students = Student.objects.all().order_by('class_name', 'name')
        
        row = 2
        for student in students:
            # Get today's attendance for the student
            today = timezone.now().date()
            attendance_records = Attendance.objects.filter(
                student=student,
                date=today
            ).order_by('time_in')

            if attendance_records.exists():
                for record in attendance_records:
                    ws.cell(row=row, column=1, value=student.lrn)
                    ws.cell(row=row, column=2, value=student.name)
                    ws.cell(row=row, column=3, value=student.class_name)
                    ws.cell(row=row, column=4, value=record.date.strftime('%Y-%m-%d'))
                    ws.cell(row=row, column=5, value=record.time_in.strftime('%H:%M:%S') if record.time_in else '')
                    ws.cell(row=row, column=6, value=record.time_out.strftime('%H:%M:%S') if record.time_out else '')
                    ws.cell(row=row, column=7, value=record.status)
                    row += 1
            else:
                # Include students with no attendance record
                ws.cell(row=row, column=1, value=student.lrn)
                ws.cell(row=row, column=2, value=student.name)
                ws.cell(row=row, column=3, value=student.class_name)
                ws.cell(row=row, column=4, value=today.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=5, value='ABSENT')
                row += 1

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Create the response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=attendance_report_{today.strftime("%Y%m%d")}.xlsx'

        # Save the workbook to the response
        wb.save(response)
        return response

    except Exception as e:
        return HttpResponse(f"Error generating Excel file: {str(e)}", status=500)